#-*- coding:utf-8 -*-
import re
import sys
import os
# import requests
import csv
import numpy as np
import matplotlib.pyplot as plt
from sklearn import linear_model
from openpyxl import Workbook
from scipy.interpolate import interp1d

def getID():
    dr = 'wdzj_feature'
       
    meanlist=[]

    with open('feature_all.txt', 'r') as f:
        content = f.read().strip()
    print '_________get platform ID'

    pattern=re.compile('.*?Entries(.*?)18 前五十投资人待收占比.*?',re.S)
    plats = re.findall(pattern,content)
    count=2
    idlist=[]
    for plat in plats:
        pattern_name=re.compile('PlatId=(.*?) .*?\n')
        name=re.findall(pattern_name,plat)
        idlist.append(name[0])
    return idlist
    
def getIDdic(idlist):
    dic={}
    for i in xrange(len(idlist)):
        dic[idlist[i]]=i
    return dic

def getitem(index):
    
    dr = 'wdzj_feature'
    title=['成交量','利率','当日待还','当日待还30日平均','资金净流入','投资人数','借款人数','人均投资金额','人均借款金额','借款标数','平均借款期限','待收投资人数','待还借款人数','投资人HHI','借款人HHI','未来60日资金流出走势','未来每日还款金额']
       
    meanlist=[]

    with open('feature_all.txt', 'r') as f:
        content = f.read().strip()
    print '_________get feature:', title[index]

    pattern=re.compile('.*?Entries(.*?)18 前五十投资人待收占比.*?',re.S)
    
    plats = re.findall(pattern,content)
    count=2
    idlist=[]
    for plat in plats:
        pattern_name=re.compile('PlatId=(.*?) .*?\n')
        name=re.findall(pattern_name,plat)
        
        idlist.append(name[0])

        pattern_content=re.compile('\[(.*?)\]',re.S)
        items=re.findall(pattern_content,plat)
        

        
        itemlist= items[index].split(",") 
        
            
        if len(itemlist)<365:
           
            for i in xrange(365-len(itemlist)):
                meanlist.append(0.0)
                
        for i in xrange(len(itemlist)):
            
            meanlist.append(float(itemlist[i]))
            
        count+=1
        

    item_matrix=np.array((meanlist)).reshape(count-2,365)#row,column
    
    return item_matrix


def getFlatMean(matrix):
    # item_mean1=matrix.mean(0)
    item_mean1=np.average(matrix,weights=(matrix!=0),axis=0)
    item_mean=item_mean1
    for j in xrange(365):
            # print mean[i][0:7].mean()
            if j <= 7:
                item_mean[j]=item_mean1[0:j+7].mean()
            else:
                if j>=365-7:
                    item_mean[j]=item_mean1[j-7:365].mean()
                else:
                    item_mean[j]=item_mean1[j-7:j+7].mean()
    return item_mean

def writeMean(sheet,item_mean,count):
    sheet.cell(row=count,column =1).value="Mean"
    for i in xrange(len(item_mean)):
        sheet.cell(row=count,column =i+2).value=float(item_mean[i])

def getRate(matrix,mean):
    plat_num=len(matrix)
    rate_matrix=np.empty([plat_num,365])
    for i in xrange(plat_num):
        for j in xrange(365):
            rate_matrix[i][j]=matrix[i][j]/mean[j]
    return rate_matrix

def getMinus(matrix,mean):
    plat_num=len(matrix)
    minus_matrix=np.empty([plat_num,365])
    for i in xrange(plat_num):
        for j in xrange(365):
            minus_matrix[i][j]=matrix[i][j]-mean[j]
    return minus_matrix

def getFlat(matrix):
    plat_num=len(matrix)
    flat_matrix=np.empty([plat_num,365])
    for i in xrange(plat_num):
        for j in xrange(365):
            # print mean[i][0:7].mean()
            if j <= 7:
                flat_matrix[i][j]=matrix[i][0:j+7].mean()
            else:
                if j>=365-7:
                    flat_matrix[i][j]=matrix[i][j-7:365].mean()
                else:
                    flat_matrix[i][j]=matrix[i][j-7:j+7].mean()
    return flat_matrix

def getSlope(matrix):
    lr=linear_model.LinearRegression()
    plat_num=len(matrix)
    slope_matrix=np.empty([plat_num,365])
    for i in xrange(plat_num):
        list_tmp=matrix[i]
        for j in xrange(365):
            if j <= 7:
                value=lr.fit(np.array(xrange(7)).reshape(-1,1),np.array((list_tmp[0:7])).reshape(-1,1),sample_weight=None)
                # print value.coef_ 
                slope_matrix[i][j]=value.coef_[0][0]

                
            else:
                if j>=365-7:
                    # print len(list_tmp)
                    # print np.array((list_tmp[365-7:365]))
                    value=lr.fit(np.array(xrange(7)).reshape(-1,1),np.array((list_tmp[365-7:365])).reshape(-1,1),sample_weight=None)
                    # print value.coef_ 
                    slope_matrix[i][j]=value.coef_[0][0]
                else:

                    value=lr.fit(np.array(xrange(len(list_tmp[j-7:j+7]))).reshape(-1,1),np.array((list_tmp[j-7:j+7])).reshape(-1,1),sample_weight=None)
                    # print value.coef_ ,j 
                    slope_matrix[i][j]=value.coef_[0][0]
    return slope_matrix

def getMapping(matrix):
    min=matrix.min()
    max=matrix.max()
    m = interp1d([min,max],[0,1])
    mapping_matrix=m(matrix)
    return mapping_matrix

def writeMatrix(sheet,idlist,matrix):
    plat_num=len(matrix)
    sheet.append(["ID"])
    for i in xrange(plat_num):
        sheet.cell(row=i+2,column=1).value=idlist[i]
    for i in xrange(plat_num):
        
        for j in xrange(365):
            sheet.cell(row=i+2,column=j+2).value=matrix[i][j]

def constructFeature(matrix,idlist,path):
    count=len(idlist)+2
    print count
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1,column=1).value='ID'
    flatmean=getFlatMean(matrix)
    writeMatrix(sheet,idlist,matrix)
    writeMean(sheet,flatmean,count)
    # mean=np.average(matrix,weights=(matrix!=0),axis=0)
    mean=flatmean
    sheet2 = wb.create_sheet(title="rate")
    rate_matrix=getRate(matrix,mean)
    writeMatrix(sheet2,idlist,rate_matrix)
    # print mean
    # print matrix
    sheet3=wb.create_sheet(title="normalize")
    flat_matrix=getFlat(matrix)
    # print flat_matrix
    writeMatrix(sheet3,idlist,flat_matrix)

    sheet4=wb.create_sheet(title="slop")
    slope_matrix=getSlope(flat_matrix)
    writeMatrix(sheet4,idlist,slope_matrix)

    sheet5= wb.create_sheet(title="minus")
    minus_matrix=getMinus(matrix,mean)
    writeMatrix(sheet5,idlist,minus_matrix)

    wb.save(path+'.xlsx')   

def constructMatrix(matrix,index):
    # idlist=getID()
    # iddic=getIDdic(idlist)
    flatmean=getFlatMean(matrix)
    mean=flatmean
    rate_matrix=getMapping(getRate(matrix,mean))
    rate_featrue=rate_matrix[index]

    flat_matrix=getMapping(getFlat(matrix))
    flat_feature=flat_matrix[index]

    slope_matrix=getMapping(getSlope(flat_matrix))
    slope_feature=slope_matrix[index]

    minus_matrix=getMapping(getMinus(matrix,mean))
    minus_feature=minus_matrix[index]

    M = np.array([])
    M=np.concatenate((M,rate_featrue))
    M=np.vstack((M,flat_feature))
    M=np.vstack((M,slope_feature))
    M=np.vstack((M,minus_feature))
    return M


def getPlatMatrix(PlatId):
    '''
    title=['成交量','利率','当日待还','当日待还30日平均','资金净流入','投资人数','借款人数','人均投资金额','人均借款金额','借款标数','平均借款期限','待收投资人数','待还借款人数','投资人HHI','借款人HHI','未来60日资金流出走势','未来每日还款金额']
    plt.plot(mapping_matrix[0])
    plt.show()
    constructFeature(matrix,idlist,'average_instrument')
    '''

    
    
    idlist=getID()
    iddic=getIDdic(idlist)
    plat=iddic[PlatId]

    matrix=getitem(0)
    M=constructMatrix(matrix,plat)

    matrix=getitem(1)
    M1=constructMatrix(matrix,plat)
    M=np.concatenate((M,M1))


    matrix0=getitem(6)#plat,365 matrix[i][j] ithplat jth day data
    matrix1=getitem(9)
    matrix=np.nan_to_num(matrix1/matrix0)
    M2=constructMatrix(matrix,plat)
    M=np.vstack((M,M2))

    print M.shape

    return M


M=getPlatMatrix('34')
# print M
# print M.shape


